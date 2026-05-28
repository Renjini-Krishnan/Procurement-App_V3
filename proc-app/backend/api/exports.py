"""Export endpoints — PPT, Excel, CSV, JSON; comparison view."""
from __future__ import annotations

import csv
import io
import json
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from .. import db
from ..engine import orchestrator
from ..services import ppt_generator, upload_service


router = APIRouter(prefix="/api/engagement", tags=["exports"])


class ExportRequest(BaseModel):
    upload_id: str
    industry: str = "steel"


def _slug(s: str) -> str:
    return "".join(c if c.isalnum() else "-" for c in s).strip("-").lower() or "engagement"


# ============================================================================
# PPT decks
# ============================================================================

@router.post("/{engagement_id}/export/findings-deck.pptx")
def export_findings_deck(engagement_id: str, payload: ExportRequest):
    eng = db.get_engagement(engagement_id)
    if not eng:
        raise HTTPException(404, "Engagement not found")
    try:
        kp = orchestrator.run_kpi_dashboard(
            engagement_id=engagement_id, upload_id=payload.upload_id,
            industry=payload.industry,
        )
        pptx_bytes = ppt_generator.generate_findings_deck(eng, kp)
    except Exception as e:
        raise HTTPException(500, f"PPT generation failed: {e}")
    name = f"procvault-findings-{_slug(eng['client_name'])}-{datetime.now().strftime('%Y%m%d')}.pptx"
    return StreamingResponse(
        io.BytesIO(pptx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )


@router.post("/{engagement_id}/export/exec-summary.pptx")
def export_exec_summary(engagement_id: str, payload: ExportRequest):
    eng = db.get_engagement(engagement_id)
    if not eng:
        raise HTTPException(404, "Engagement not found")
    try:
        kp = orchestrator.run_kpi_dashboard(
            engagement_id=engagement_id, upload_id=payload.upload_id,
            industry=payload.industry,
        )
        pptx_bytes = ppt_generator.generate_exec_summary_deck(eng, kp)
    except Exception as e:
        raise HTTPException(500, f"PPT generation failed: {e}")
    name = f"procvault-exec-summary-{_slug(eng['client_name'])}-{datetime.now().strftime('%Y%m%d')}.pptx"
    return StreamingResponse(
        io.BytesIO(pptx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )


# ============================================================================
# Excel export — KPIs across all pillars
# ============================================================================

@router.post("/{engagement_id}/export/kpis.xlsx")
def export_kpis_xlsx(engagement_id: str, payload: ExportRequest):
    eng = db.get_engagement(engagement_id)
    if not eng:
        raise HTTPException(404, "Engagement not found")
    try:
        import openpyxl
    except Exception:
        raise HTTPException(500, "openpyxl not installed")

    kp = orchestrator.run_kpi_dashboard(engagement_id, payload.upload_id, payload.industry)

    wb = openpyxl.Workbook()

    # Sheet 1: KPIs
    ws = wb.active
    ws.title = "KPIs"
    cols = ["ID", "Label", "Pillar", "Theme", "Value", "Unit", "Band low", "Band high",
            "Band meaning", "Status", "Delta",
            "Benchmark source", "Benchmark year", "Confidence", "Finding"]
    ws.append(cols)
    for k in kp.get("kpis", []):
        ws.append([
            k.get("id"), k.get("label"), k.get("pillar"), k.get("theme"),
            k.get("value"), k.get("unit"),
            k.get("band", {}).get("low"), k.get("band", {}).get("high"),
            k.get("band_meaning"), k.get("status"), k.get("delta"),
            k.get("benchmark", {}).get("source"),
            k.get("benchmark", {}).get("year"),
            k.get("benchmark", {}).get("confidence"),
            k.get("finding"),
        ])

    # Sheet 2: Pillar summary
    ws2 = wb.create_sheet("Pillar summary")
    ws2.append(["Pillar", "Score", "Label", "KPI count", "In band", "Below", "Above"])
    for pid, s in kp.get("pillar_summary", {}).items():
        ws2.append([
            pid,
            (s.get("pillar_score") or {}).get("score"),
            (s.get("pillar_score") or {}).get("label"),
            s.get("kpi_count"), s.get("in_band"), s.get("under"), s.get("over"),
        ])

    # Sheet 3: Findings
    ws3 = wb.create_sheet("Findings")
    ws3.append(["Pillar", "Theme", "Headline"])
    for f in orchestrator.get_findings(engagement_id):
        ws3.append([f.get("pillar"), f.get("theme"), f.get("headline")])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    name = f"procvault-kpis-{_slug(eng['client_name'])}-{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )


# ============================================================================
# Comparison — current vs prior pillar run
# ============================================================================

@router.get("/{engagement_id}/comparison")
def compare_runs(engagement_id: str, pillar: Optional[str] = None):
    """Return current + previous pillar runs with deltas + KPI-level
    differences if cached run output is available."""
    if not db.get_engagement(engagement_id):
        raise HTTPException(404, "Engagement not found")

    runs = db.list_pillar_runs(engagement_id, pillar=pillar, limit=200)
    if not runs:
        return {"engagement_id": engagement_id, "comparisons": []}

    # Group by pillar; sort newest→oldest
    by_pillar: dict[str, list[dict]] = {}
    for r in runs:
        by_pillar.setdefault(r["pillar"], []).append(r)
    for p in by_pillar:
        by_pillar[p].sort(key=lambda r: r["ran_at"], reverse=True)

    comparisons = []
    for p, rs in by_pillar.items():
        current = rs[0]
        prior = rs[1] if len(rs) > 1 else None
        cur_score = current.get("pillar_score")
        prior_score = prior.get("pillar_score") if prior else None
        delta = (cur_score - prior_score) if (cur_score is not None and prior_score is not None) else None
        # Theme-level deltas
        theme_deltas = []
        if prior:
            cur_ts = current.get("theme_scores") or {}
            prior_ts = prior.get("theme_scores") or {}
            for tid, cv in cur_ts.items():
                pv = prior_ts.get(tid)
                if isinstance(cv, (int, float)) and isinstance(pv, (int, float)):
                    theme_deltas.append({
                        "theme": tid, "current": cv, "prior": pv, "delta": cv - pv,
                    })
        comparisons.append({
            "pillar": p,
            "current": {
                "ran_at": current["ran_at"],
                "score": cur_score,
                "label": current.get("pillar_label"),
                "headline": current.get("headline", ""),
            },
            "prior": {
                "ran_at": prior["ran_at"],
                "score": prior_score,
                "label": prior.get("pillar_label"),
                "headline": prior.get("headline", ""),
            } if prior else None,
            "delta": delta,
            "trend": "up" if (delta or 0) > 0 else "down" if (delta or 0) < 0 else "flat",
            "run_count": len(rs),
            "theme_deltas": theme_deltas,
            "history": [
                {"ran_at": r["ran_at"], "score": r["pillar_score"]}
                for r in rs[:20]
            ],
        })

    return {"engagement_id": engagement_id, "comparisons": comparisons}
