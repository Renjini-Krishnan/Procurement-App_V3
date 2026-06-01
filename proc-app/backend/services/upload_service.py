"""Upload + parse service.

Handles client file uploads (PO dumps initially), persists to filesystem,
records metadata in SQLite, and surfaces the raw columns for Stage 5/6
column-mapping HITL.
"""
from __future__ import annotations

import io
import json
import shutil
import uuid
from pathlib import Path
from typing import Any, Optional

import pandas as pd

from .. import config, db
from . import canonical_schema


UPLOADS_DIR = Path(__file__).resolve().parents[1] / "data" / "uploads"
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

import hashlib
import os

MAX_UPLOAD_SIZE_MB = int(os.environ.get("PROCVAULT_MAX_UPLOAD_MB", "100"))
MAX_UPLOAD_SIZE_BYTES = MAX_UPLOAD_SIZE_MB * 1024 * 1024


class FileTooLargeError(ValueError):
    pass


class DuplicateUploadError(ValueError):
    def __init__(self, existing_upload_id: str, *args):
        super().__init__(*args)
        self.existing_upload_id = existing_upload_id


def save_upload(
    engagement_id: str,
    file_type: str,
    original_filename: str,
    file_bytes: bytes,
    *,
    skip_dedup: bool = False,
    auto_classified: bool = False,
) -> dict:
    """Persist uploaded file to disk + insert into SQLite uploads table.

    Raises:
      FileTooLargeError if bytes > MAX_UPLOAD_SIZE_BYTES.
      DuplicateUploadError if a prior upload in the same engagement has the
      same SHA256 hash (skip_dedup=True bypasses).
    """
    size_bytes = len(file_bytes)
    if size_bytes > MAX_UPLOAD_SIZE_BYTES:
        raise FileTooLargeError(
            f"File is {size_bytes / 1024 / 1024:.1f} MB; limit is "
            f"{MAX_UPLOAD_SIZE_MB} MB. Set PROCVAULT_MAX_UPLOAD_MB to raise."
        )

    content_hash = hashlib.sha256(file_bytes).hexdigest()
    if not skip_dedup:
        existing = _find_dup(engagement_id, content_hash)
        if existing:
            raise DuplicateUploadError(
                existing["id"],
                f"Same file already uploaded (id={existing['id']}, "
                f"file_type={existing['file_type']}, "
                f"uploaded {existing['uploaded_at']}).",
            )

    upload_id = uuid.uuid4().hex[:12]
    engagement_dir = UPLOADS_DIR / engagement_id
    engagement_dir.mkdir(parents=True, exist_ok=True)
    suffix = Path(original_filename).suffix.lower() or ".csv"
    stored_path = engagement_dir / f"{upload_id}{suffix}"
    stored_path.write_bytes(file_bytes)

    # Parse — CSV or Excel
    try:
        if suffix in (".csv", ".tsv", ".txt"):
            df = pd.read_csv(io.BytesIO(file_bytes), low_memory=False)
        else:
            df = pd.read_excel(io.BytesIO(file_bytes))
    except Exception as e:
        # Clean up file + raise
        stored_path.unlink(missing_ok=True)
        raise ValueError(f"Could not parse file '{original_filename}': {e}") from e

    raw_columns = [str(c).strip() for c in df.columns]
    row_count = len(df)
    sample_rows = df.head(5).fillna("").astype(str).values.tolist()

    # Suggest column mapping
    suggestion = canonical_schema.suggest_mapping(raw_columns, file_type)

    # Insert into uploads table
    ts = db.now_iso()
    with db.db_connection() as conn:
        conn.execute(
            """
            INSERT INTO uploads
            (id, engagement_id, file_type, original_filename, stored_path,
             row_count, column_mapping, content_hash, size_bytes,
             auto_classified, uploaded_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                upload_id, engagement_id, file_type, original_filename,
                str(stored_path), row_count,
                json.dumps({"suggested": suggestion["matches"], "confirmed": None}),
                content_hash, size_bytes, 1 if auto_classified else 0, ts,
            ),
        )

    return {
        "upload_id": upload_id,
        "engagement_id": engagement_id,
        "file_type": file_type,
        "original_filename": original_filename,
        "row_count": row_count,
        "columns": raw_columns,
        "sample_rows": sample_rows,
        "suggested_mapping": suggestion["matches"],
        "missing_required": suggestion["missing_required"],
        "schema": suggestion["schema"],
        "uploaded_at": ts,
    }


def get_upload(upload_id: str) -> Optional[dict]:
    with db.db_connection() as conn:
        row = conn.execute("SELECT * FROM uploads WHERE id = ?", (upload_id,)).fetchone()
    if not row:
        return None
    d = dict(row)
    if d.get("column_mapping"):
        try:
            d["column_mapping"] = json.loads(d["column_mapping"])
        except Exception:
            pass
    return d


def list_uploads(engagement_id: str) -> list[dict]:
    with db.db_connection() as conn:
        rows = conn.execute(
            "SELECT * FROM uploads WHERE engagement_id = ? ORDER BY uploaded_at DESC",
            (engagement_id,),
        ).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        if d.get("column_mapping"):
            try:
                d["column_mapping"] = json.loads(d["column_mapping"])
            except Exception:
                pass
        out.append(d)
    return out


def confirm_mapping(upload_id: str, confirmed_mapping: list[dict]) -> dict:
    """Persist consultant-confirmed column mapping.

    confirmed_mapping = [
      {"raw_column": "Pur.Order Number", "canonical_field": "po_number"},
      ...
    ]
    """
    upload = get_upload(upload_id)
    if not upload:
        raise ValueError(f"Upload {upload_id} not found")

    schema = canonical_schema.get_schema(upload["file_type"])
    required_fields = {f["field"] for f in schema["fields"] if f["required"]}
    mapped_canonical = {m["canonical_field"] for m in confirmed_mapping if m.get("canonical_field")}
    missing = sorted(required_fields - mapped_canonical)

    # Merge: keep suggested + store confirmed
    current = upload.get("column_mapping") or {}
    if not isinstance(current, dict):
        current = {}
    current["confirmed"] = confirmed_mapping
    current["missing_required"] = missing

    with db.db_connection() as conn:
        conn.execute(
            "UPDATE uploads SET column_mapping = ? WHERE id = ?",
            (json.dumps(current), upload_id),
        )

    return {
        "upload_id": upload_id,
        "confirmed_mapping": confirmed_mapping,
        "missing_required": missing,
        "ready_for_bronze": len(missing) == 0,
    }


def list_validation_status(engagement_id: str) -> dict:
    """Per-engagement view: which uploads still need column mapping confirmed.
    Used by Stage 6 to drive the multi-file tracker + only advance to Bronze
    once every upload is mapped."""
    uploads = list_uploads(engagement_id)
    items = []
    confirmed_count = 0
    for u in uploads:
        cm = u.get("column_mapping") or {}
        if not isinstance(cm, dict):
            cm = {}
        confirmed = cm.get("confirmed")
        missing = cm.get("missing_required") or []
        is_confirmed = bool(confirmed) and len(missing) == 0
        if is_confirmed:
            confirmed_count += 1
        items.append({
            "upload_id": u["id"],
            "file_type": u.get("file_type"),
            "original_filename": u.get("original_filename"),
            "row_count": u.get("row_count"),
            "is_confirmed": is_confirmed,
            "missing_required": missing,
        })
    return {
        "engagement_id": engagement_id,
        "total": len(items),
        "confirmed": confirmed_count,
        "all_ready_for_bronze": (len(items) > 0 and confirmed_count == len(items)),
        "uploads": items,
    }


def read_upload_dataframe(upload_id: str) -> pd.DataFrame:
    """Read the stored file back into a DataFrame for downstream stages."""
    upload = get_upload(upload_id)
    if not upload:
        raise ValueError(f"Upload {upload_id} not found")
    path = Path(upload["stored_path"])
    suffix = path.suffix.lower()
    if suffix in (".csv", ".tsv", ".txt"):
        return pd.read_csv(path, low_memory=False)
    return pd.read_excel(path)


SEED_DIR = Path(__file__).resolve().parents[1] / "data" / "seed"

SEED_FILES = {
    "PO":              "demo_po_dump.csv",
    "PR":              "demo_pr_dump.csv",
    "VENDOR_MASTER":   "demo_vendor_master.csv",
    "MATERIAL_MASTER": "demo_material_master.csv",
    "ORG_STRUCTURE":   "demo_org_structure.csv",
    "CONTRACT_MASTER": "demo_contract_master.csv",
    "GRN":             "demo_grn.csv",
    "INVOICE":         "demo_invoice.csv",
}


def get_seed_dataset_path(file_type: str = "PO") -> Path:
    name = SEED_FILES.get(file_type.upper())
    if not name:
        raise FileNotFoundError(f"No seed configured for {file_type}")
    return SEED_DIR / name


def list_available_seeds() -> list[dict]:
    out = []
    for ft, name in SEED_FILES.items():
        p = SEED_DIR / name
        if p.exists():
            out.append({"file_type": ft, "filename": name,
                         "size_bytes": p.stat().st_size,
                         "row_count_estimate": sum(1 for _ in open(p, "r", encoding="utf-8")) - 1})
    return out


def use_seed_dataset(engagement_id: str, file_type: str = "PO") -> dict:
    """Load the seed CSV for the requested file_type as if uploaded.
    Auto-confirms the suggested mapping so downstream stages (Bronze, Gold,
    KPIs) don't trip on "no mapping confirmed" — the seeds are known-good
    data with deterministic column names."""
    seed_path = get_seed_dataset_path(file_type)
    if not seed_path.exists():
        raise FileNotFoundError(f"Seed dataset not found at {seed_path}")
    result = save_upload(
        engagement_id=engagement_id,
        file_type=file_type.upper(),
        original_filename=seed_path.name,
        file_bytes=seed_path.read_bytes(),
        skip_dedup=True,  # demo seeds are allowed to be re-loaded
    )
    # Auto-confirm the suggested mapping (seeds are known-good; no HITL needed)
    try:
        confirmed = [
            {"raw_column": m["raw_column"], "canonical_field": m.get("suggested_field")}
            for m in result.get("suggested_mapping", [])
        ]
        if confirmed:
            confirm_mapping(result["upload_id"], confirmed)
    except Exception:
        # Don't fail the seed load if auto-confirm fails — user can still
        # walk through Stage 6 manually.
        pass
    return result


def load_all_seeds(engagement_id: str) -> dict:
    """Load every seed dataset (PO + PR + GRN + INVOICE + masters) in one
    call. Skips files that fail (e.g. dedup hit) and returns a per-file
    report. Used by the "Load all sample data" button in Stage 4."""
    results = []
    success = 0
    for ft in SEED_FILES.keys():
        try:
            r = use_seed_dataset(engagement_id, ft)
            results.append({
                "file_type": ft,
                "status": "loaded",
                "upload_id": r["upload_id"],
                "filename": r["original_filename"],
                "row_count": r["row_count"],
            })
            success += 1
        except DuplicateUploadError as e:
            results.append({
                "file_type": ft, "status": "duplicate",
                "existing_upload_id": e.existing_upload_id,
                "reason": str(e),
            })
        except Exception as e:
            results.append({
                "file_type": ft, "status": "failed",
                "reason": f"{type(e).__name__}: {e}",
            })
    return {
        "engagement_id": engagement_id,
        "total_seeds": len(SEED_FILES),
        "loaded": success,
        "files": results,
    }


# --------------------------------------------------------------------------
# Dedup helper
# --------------------------------------------------------------------------

def _find_dup(engagement_id: str, content_hash: str) -> Optional[dict]:
    with db.db_connection() as conn:
        row = conn.execute(
            "SELECT * FROM uploads WHERE engagement_id = ? AND content_hash = ? "
            "ORDER BY uploaded_at DESC LIMIT 1",
            (engagement_id, content_hash),
        ).fetchone()
    return dict(row) if row else None


# --------------------------------------------------------------------------
# Auto-classifier — pick best file_type by alias overlap with each schema
# --------------------------------------------------------------------------

def classify_file_type(raw_columns: list[str]) -> dict:
    """Return {best, score, scores: {file_type: pct}, confidence}.

    Compares the normalised raw header set against every canonical schema's
    alias set; picks the file_type whose required-field aliases cover the
    largest share of the raw headers.
    """
    norm_raw = {canonical_schema._normalise(c) for c in raw_columns if c}
    if not norm_raw:
        return {"best": None, "score": 0.0, "scores": {}, "confidence": "none"}

    scores = {}
    for ft, _yml in canonical_schema._FILE_NAME_BY_TYPE.items():
        try:
            schema = canonical_schema.get_schema(ft)
        except Exception:
            continue
        # Required-field aliases get full weight; optional half weight
        req_aliases, opt_aliases = set(), set()
        for f in schema["fields"]:
            target = req_aliases if f.get("required") else opt_aliases
            target.add(canonical_schema._normalise(f["field"]))
            for a in f.get("aliases", []):
                target.add(canonical_schema._normalise(a))
        req_hits = sum(1 for c in norm_raw if c in req_aliases)
        opt_hits = sum(1 for c in norm_raw if c in opt_aliases and c not in req_aliases)
        # Score = (required hits + 0.5 * optional hits) / raw column count, capped at 1.0
        scores[ft] = round(min(1.0, (req_hits + 0.5 * opt_hits) / max(len(norm_raw), 1)), 3)

    best = max(scores.items(), key=lambda x: x[1])
    confidence = ("high" if best[1] >= 0.5
                   else "medium" if best[1] >= 0.25
                   else "low" if best[1] >= 0.1
                   else "none")
    return {"best": best[0], "score": best[1], "scores": scores, "confidence": confidence}


# --------------------------------------------------------------------------
# Per-upload quick-stats — surfaced on Stage 5 AI Validation
# --------------------------------------------------------------------------

def _resolve_columns(raw_columns: list[str], file_type: str,
                      wanted_fields: list[str]) -> dict[str, Optional[str]]:
    """Best-effort: map canonical field → raw column name using suggest_mapping."""
    suggestion = canonical_schema.suggest_mapping(raw_columns, file_type)
    by_canonical: dict[str, str] = {}
    for m in suggestion.get("matches", []):
        if m.get("suggested_field") and m["suggested_field"] not in by_canonical:
            by_canonical[m["suggested_field"]] = m["raw_column"]
    return {fld: by_canonical.get(fld) for fld in wanted_fields}


def _detect_currency_unit(values: pd.Series) -> tuple[float, str]:
    """Sum-based unit detection (mirrors KB cleansing rule). Returns
    (multiplier_to_INR, label). If the magnitude suggests thousands or
    lakhs/crores in header text we'd handle that elsewhere — here we use
    a conservative sum-magnitude heuristic only as a fallback label."""
    total = float(pd.to_numeric(values, errors="coerce").fillna(0).sum())
    # Conservative: assume the column is already in INR; just return total + label.
    return total, "INR"


def _format_inr(amount: float) -> str:
    if amount is None or amount != amount:  # NaN check
        return "—"
    a = abs(amount)
    if a >= 1e7:
        return f"₹{amount/1e7:,.2f} Cr"
    if a >= 1e5:
        return f"₹{amount/1e5:,.2f} L"
    if a >= 1e3:
        return f"₹{amount/1e3:,.1f} K"
    return f"₹{amount:,.0f}"


def _date_range(series: pd.Series) -> Optional[dict]:
    try:
        s = pd.to_datetime(series, errors="coerce", dayfirst=True)
        s = s.dropna()
        if len(s) == 0:
            return None
        return {
            "min": s.min().strftime("%Y-%m-%d"),
            "max": s.max().strftime("%Y-%m-%d"),
            "span_days": int((s.max() - s.min()).days),
            "parsed_rows": int(len(s)),
        }
    except Exception:
        return None


def _quick_stats_for(file_type: str, df: pd.DataFrame) -> dict:
    """Compute file-type-specific summary stats. Best-effort; missing
    columns just yield fewer fields. NEVER raises."""
    ft = file_type.upper()
    raw_cols = [str(c).strip() for c in df.columns]
    stats: dict[str, Any] = {"row_count": int(len(df))}

    def _col(canonical: str) -> Optional[str]:
        return _resolve_columns(raw_cols, ft, [canonical]).get(canonical)

    try:
        if ft == "PO":
            amt_col = _col("net_value") or _col("gr_value") or _col("invoice_value")
            if amt_col and amt_col in df.columns:
                total, _ = _detect_currency_unit(df[amt_col])
                stats["total_value_inr"] = total
                stats["total_value_label"] = _format_inr(total)
            vendor_col = _col("vendor_id") or _col("vendor_name")
            if vendor_col and vendor_col in df.columns:
                stats["distinct_vendors"] = int(df[vendor_col].dropna().nunique())
            mg_col = _col("material_group_desc") or _col("material_group")
            if mg_col and mg_col in df.columns:
                grp = df[mg_col].dropna().astype(str)
                stats["distinct_material_groups"] = int(grp.nunique())
                if amt_col and amt_col in df.columns:
                    vals = pd.to_numeric(df[amt_col], errors="coerce").fillna(0)
                    top = (pd.DataFrame({"mg": grp, "v": vals})
                            .groupby("mg")["v"].sum().sort_values(ascending=False).head(5))
                    stats["top_categories"] = [
                        {"label": k, "value_inr": float(v), "value_label": _format_inr(float(v))}
                        for k, v in top.items()
                    ]
            date_col = _col("po_creation_date") or _col("delivery_date")
            if date_col and date_col in df.columns:
                dr = _date_range(df[date_col]);  dr and stats.update({"date_range": dr})

        elif ft == "PR":
            pr_col = _col("pr_number")
            if pr_col and pr_col in df.columns:
                stats["distinct_prs"] = int(df[pr_col].dropna().nunique())
            amt_col = _col("pr_total_value")
            if amt_col and amt_col in df.columns:
                total, _ = _detect_currency_unit(df[amt_col])
                stats["total_value_inr"] = total
                stats["total_value_label"] = _format_inr(total)
            date_col = _col("pr_creation_date") or _col("pr_release_date")
            if date_col and date_col in df.columns:
                dr = _date_range(df[date_col]);  dr and stats.update({"date_range": dr})

        elif ft == "GRN":
            grn_col = _col("grn_number")
            if grn_col and grn_col in df.columns:
                stats["distinct_grns"] = int(df[grn_col].dropna().nunique())
            amt_col = _col("amount")
            if amt_col and amt_col in df.columns:
                total, _ = _detect_currency_unit(df[amt_col])
                stats["total_value_inr"] = total
                stats["total_value_label"] = _format_inr(total)
            date_col = _col("posting_date") or _col("document_date")
            if date_col and date_col in df.columns:
                dr = _date_range(df[date_col]);  dr and stats.update({"date_range": dr})

        elif ft == "INVOICE":
            inv_col = _col("invoice_doc_number")
            if inv_col and inv_col in df.columns:
                stats["distinct_invoices"] = int(df[inv_col].dropna().nunique())
            amt_col = _col("gross_amount")
            if amt_col and amt_col in df.columns:
                total, _ = _detect_currency_unit(df[amt_col])
                stats["total_value_inr"] = total
                stats["total_value_label"] = _format_inr(total)
            date_col = _col("invoice_date") or _col("posting_date")
            if date_col and date_col in df.columns:
                dr = _date_range(df[date_col]);  dr and stats.update({"date_range": dr})

        elif ft == "VENDOR_MASTER":
            vid_col = _col("vendor_id")
            if vid_col and vid_col in df.columns:
                stats["distinct_vendors"] = int(df[vid_col].dropna().nunique())

        elif ft == "MATERIAL_MASTER":
            mid_col = _col("material_number")
            if mid_col and mid_col in df.columns:
                stats["distinct_materials"] = int(df[mid_col].dropna().nunique())

        elif ft == "ORG_STRUCTURE":
            emp_col = _col("employee_id")
            if emp_col and emp_col in df.columns:
                stats["distinct_employees"] = int(df[emp_col].dropna().nunique())

        elif ft == "CONTRACT_MASTER":
            c_col = _col("contract_number")
            if c_col and c_col in df.columns:
                stats["distinct_contracts"] = int(df[c_col].dropna().nunique())
    except Exception as e:
        stats["_stats_error"] = str(e)

    return stats


def build_upload_summary(engagement_id: str) -> dict:
    """Return per-upload roster: classification confidence (recomputed) +
    quick stats. Used by Stage 5 AI Validation."""
    rows = list_uploads(engagement_id)
    uploads_summary = []
    for u in rows:
        try:
            df = read_upload_dataframe(u["id"])
        except Exception as e:
            uploads_summary.append({
                "upload_id": u["id"],
                "file_type": u.get("file_type"),
                "original_filename": u.get("original_filename"),
                "row_count": u.get("row_count"),
                "uploaded_at": u.get("uploaded_at"),
                "auto_classified": bool(u.get("auto_classified")),
                "_error": f"Could not read file: {e}",
            })
            continue
        raw_cols = [str(c).strip() for c in df.columns]
        classification = classify_file_type(raw_cols)
        # Top 3 alternatives (excluding the best) by score
        alternatives = sorted(
            [(k, v) for k, v in classification["scores"].items() if k != classification["best"]],
            key=lambda x: -x[1],
        )[:3]
        ft_used = u.get("file_type") or classification["best"]
        quick = _quick_stats_for(ft_used, df) if ft_used else {"row_count": int(len(df))}
        classification_matches_persisted = (classification["best"] == u.get("file_type"))
        uploads_summary.append({
            "upload_id": u["id"],
            "file_type": u.get("file_type"),
            "original_filename": u.get("original_filename"),
            "row_count": u.get("row_count"),
            "uploaded_at": u.get("uploaded_at"),
            "auto_classified": bool(u.get("auto_classified")),
            "classification": {
                "best": classification["best"],
                "score": classification["score"],
                "confidence": classification["confidence"],
                "matches_persisted_type": classification_matches_persisted,
                "alternatives": [{"file_type": k, "score": v} for k, v in alternatives],
            },
            "quick_stats": quick,
        })
    return {"engagement_id": engagement_id, "uploads": uploads_summary}


# --------------------------------------------------------------------------
# Blank-template generators (XLSX + CSV) for client distribution
# --------------------------------------------------------------------------

def blank_template_csv(file_type: str) -> bytes:
    """CSV: header row of canonical aliases + 1 empty sample row."""
    schema = canonical_schema.get_schema(file_type.upper())
    headers = []
    for f in schema["fields"]:
        # Prefer the first alias (client-friendly form), else the canonical name
        aliases = f.get("aliases") or []
        headers.append(aliases[0] if aliases else f["field"])
    import io
    import csv
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    w.writerow([""] * len(headers))   # placeholder row for the client to fill
    return buf.getvalue().encode("utf-8")


def blank_template_xlsx(file_type: str) -> bytes:
    """XLSX: 3-sheet workbook (Data · Field guide · Submission guidelines)."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    schema = canonical_schema.get_schema(file_type.upper())
    guidelines = schema.get("guidelines", {})

    wb = openpyxl.Workbook()

    # ---- Sheet 1: Data
    ws = wb.active
    ws.title = "Data"
    headers = []
    for f in schema["fields"]:
        aliases = f.get("aliases") or []
        headers.append(aliases[0] if aliases else f["field"])
    ws.append(headers)
    header_fill = PatternFill(start_color="FF1E1B4B", end_color="FF1E1B4B", fill_type="solid")
    bold_white = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = bold_white
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    # Required fields get a light highlight in the second row to flag them
    req_fill = PatternFill(start_color="FFFFF7CC", end_color="FFFFF7CC", fill_type="solid")
    for col_idx, f in enumerate(schema["fields"], start=1):
        if f.get("required"):
            ws.cell(row=2, column=col_idx).fill = req_fill
    ws.cell(row=2, column=1).comment = openpyxl.comments.Comment(
        "Required fields are highlighted yellow. Add your rows below row 1 — header should match exactly.",
        "Procvault",
    )
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22

    # ---- Sheet 2: Field guide
    ws2 = wb.create_sheet("Field guide")
    ws2.append(["#", "Canonical field", "Header to use", "Required?", "Type",
                 "Description", "Common aliases (acceptable headers)"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = bold_white
    for i, f in enumerate(schema["fields"], start=1):
        aliases = f.get("aliases") or []
        ws2.append([
            i, f["field"], aliases[0] if aliases else f["field"],
            "YES" if f.get("required") else "no",
            f.get("type", "string"), (f.get("description") or "")[:300],
            ", ".join(aliases[:8]),
        ])
    for col_letter, width in zip("ABCDEFG", [4, 24, 24, 10, 10, 60, 60]):
        ws2.column_dimensions[col_letter].width = width

    # ---- Sheet 3: Submission guidelines
    ws3 = wb.create_sheet("Submission guidelines")
    ws3.append(["Item", "Value / note"])
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = bold_white
    rows = [
        ["File type", schema.get("file_type", "")],
        ["Schema description", (schema.get("description") or "").strip()],
        ["Lookback recommended (months)", guidelines.get("lookback_months_recommended") or "n/a"],
        ["Minimum lookback required (months)", guidelines.get("minimum_months_required") or "n/a"],
        ["Date format expected", guidelines.get("date_format_expected") or "dd-mm-yyyy"],
        ["File encoding", guidelines.get("encoding_expected") or "UTF-8"],
        ["Accepted formats", ", ".join(guidelines.get("accepted_formats") or ["csv", "xlsx"])],
        ["Row granularity", guidelines.get("row_granularity") or ""],
        ["Currency handling", guidelines.get("currency_handling") or "Engine converts to INR using shared-kb FX table"],
    ]
    for r in rows:
        ws3.append(r)
    ws3.append([])
    ws3.append(["Notes for the client:"])
    ws3.cell(row=ws3.max_row, column=1).font = Font(bold=True)
    for note in (guidelines.get("notes_for_client") or []):
        ws3.append(["", "• " + note])
    for col, w in zip("AB", [38, 90]):
        ws3.column_dimensions[col].width = w
    for row in ws3.iter_rows():
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def combined_sample_xlsx_all_tabs() -> bytes:
    """Return ONE Excel workbook with ALL 8 sample datasets as separate tabs.
    Useful for a single-file demo download: clients receive one file, drop
    it in Stage 4 for any sheet, and re-upload separately for the others.
    Each tab is the seed CSV converted to a worksheet."""
    import io as _io
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill(start_color="1A1F36", end_color="1A1F36", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Cover sheet
    cover = wb.create_sheet("README")
    cover_lines = [
        ["Procvault — Combined Sample Data"],
        [""],
        ["This workbook contains 8 procurement data templates as separate tabs."],
        ["Each tab corresponds to one file_type the app expects."],
        [""],
        ["Tab", "File type", "Rows"],
    ]
    for r_idx, row in enumerate(cover_lines, start=1):
        for c_idx, val in enumerate(row, start=1):
            cover.cell(row=r_idx, column=c_idx, value=val)
    cover["A1"].font = Font(bold=True, size=14)

    file_types = ["PO", "PR", "GRN", "INVOICE", "VENDOR_MASTER",
                   "MATERIAL_MASTER", "ORG_STRUCTURE", "CONTRACT_MASTER"]
    cover_row = len(cover_lines) + 1
    for ft in file_types:
        try:
            path = get_seed_dataset_path(ft)
            if not path.exists():
                cover.cell(row=cover_row, column=1, value=ft)
                cover.cell(row=cover_row, column=2, value=ft)
                cover.cell(row=cover_row, column=3, value="(missing)")
                cover_row += 1
                continue
            df = pd.read_csv(path, low_memory=False)
            sheet_name = ft[:31]  # Excel max sheet name = 31 chars
            ws = wb.create_sheet(sheet_name)
            # Headers
            for c_idx, col in enumerate(df.columns, start=1):
                cell = ws.cell(row=1, column=c_idx, value=str(col))
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = 18
            # Data — cap at 5000 rows per sheet to keep file size reasonable
            for r_idx, row in enumerate(df.head(5000).itertuples(index=False), start=2):
                for c_idx, val in enumerate(row, start=1):
                    if pd.isna(val): val = ""
                    ws.cell(row=r_idx, column=c_idx, value=val)
            ws.freeze_panes = "A2"
            # Cover row
            cover.cell(row=cover_row, column=1, value=sheet_name)
            cover.cell(row=cover_row, column=2, value=ft)
            cover.cell(row=cover_row, column=3, value=min(len(df), 5000))
            cover_row += 1
        except Exception as e:
            cover.cell(row=cover_row, column=1, value=ft)
            cover.cell(row=cover_row, column=3, value=f"error: {e}")
            cover_row += 1
    cover.column_dimensions["A"].width = 20
    cover.column_dimensions["B"].width = 22
    cover.column_dimensions["C"].width = 14

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
